from flask import Flask, render_template, request, flash, redirect, url_for
import pandas as pd
import json
import openpyxl
import re
import xlwings as xw
import os
import psutil
import webview
import sys

app = Flask(__name__, static_folder='static', template_folder='templates')


# Utility function for handling resource paths in bundled apps
def resource_path(relative_path):
    """ Get the absolute path to the resource, works for both dev and PyInstaller """
    try:
        base_path = sys._MEIPASS  # Path where the .exe is unpacked in PyInstaller
    except AttributeError:
        base_path = os.path.abspath(".")  # Fallback for normal script execution

    return os.path.join(base_path, relative_path)


def force_close_excel():
    """Iterates over all running processes and kills any Excel instances."""
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] == 'EXCEL.EXE':
            try:
                proc.kill()  # Kill the process
                print(f"Closed Excel process with PID: {proc.info['pid']}")
            except Exception as e:
                print(f"Could not close Excel process: {e}")


# Calculation with the user input and Excel updating
def calculate_outcome(user_input):
    try:
        # Step 1: Parse the user input to extract metrics, criteria, and targets
        metrics_data = {}
        targets_data = {}

        for key, value in user_input.items():
            if key.startswith('criteria_'):
                pattern = r'criteria_(.*?)_(.*)'  # Regex to match the metric and criteria
                match = re.match(pattern, key)

                if match:
                    metric_name = match.group(1)
                    criteria_name = match.group(2)

                    if metric_name not in metrics_data:
                        metrics_data[metric_name] = [[], []]  # First list for criteria names, second for values

                    metrics_data[metric_name][0].append(criteria_name)
                    metrics_data[metric_name][1].append(value)

            elif key.startswith('target_'):
                metric_name = key.replace('target_', '')
                targets_data[metric_name] = value

        # Step 2: Load the Excel workbook and update the dataset with the extracted criteria values
        file_path = resource_path('Master health check icsa Feb 24.xlsx')
        workbook = openpyxl.load_workbook(file_path, data_only=False)  # Load to keep formulas intact
        sheet = workbook['metrics']  # Adjust to the correct sheet name if necessary

        metric_row_mapping = {}

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            metric_name = row[0].value  # Assuming metric name is in column A
            if not metric_name:
                continue  # Skip if there's no metric name

            if metric_name in metrics_data:
                criteria_list, value_list = metrics_data[metric_name]
                metric_row_mapping[metric_name] = row[0].row  # Store the row number of this metric

                for idx, criteria in enumerate(criteria_list):
                    for col_idx in range(3, 8):  # Columns D to H contain the criteria names
                        if row[col_idx].value == criteria:
                            sheet.cell(row=row[0].row, column=col_idx + 1).value = value_list[idx]
                            break

        # Save the updated workbook temporarily to recalculate outcomes
        updated_file_path = resource_path('Updated_Metrics_Dataset.xlsx')
        workbook.save(updated_file_path)

        # Step 3: Use xlwings to reopen the updated Excel file and fetch the computed outcomes
        app_xlwings = xw.App(visible=False)
        workbook_xlwings = xw.Book(updated_file_path)
        sheet_xlwings = workbook_xlwings.sheets['metrics']  # Adjust to the correct sheet name if necessary

        # Step 4: Create a new workbook for the structured output
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "Selected Metrics"

        headers = ["Metric", "Criteria", "Value", "Outcome", "Target"]
        new_sheet.append(headers)

        for metric_name, (criteria_list, value_list) in metrics_data.items():
            metric_row_in_sheet = metric_row_mapping[metric_name]
            outcome_cell = sheet_xlwings.range(f'I{metric_row_in_sheet}').value  # Assuming outcome is in column I
            target_value = targets_data.get(metric_name, 'NA')

            for idx, criteria_name in enumerate(criteria_list):
                if idx == 0:
                    new_sheet.append([metric_name, criteria_name, value_list[idx], outcome_cell, target_value])
                else:
                    new_sheet.append(["", criteria_name, value_list[idx], "", ""])

        # Step 5: Close the xlwings workbook and save the new workbook
        workbook_xlwings.close()
        app_xlwings.quit()

        new_file_path = resource_path('Structured_Metrics_Output.xlsx')
        new_workbook.save(new_file_path)

        if os.path.exists(updated_file_path):
            os.remove(updated_file_path)

        print(f"New structured Excel file saved as {new_file_path} and temporary file deleted.")

    except Exception as e:
        print(f"Error during calculation: {e}")


# Load metrics from Excel file
def load_metrics_from_excel(file_path):
    df = pd.read_excel(file_path)
    metrics_list = {}

    for index, row in df.iterrows():
        category = row[0]
        metrics = row[1:].dropna().tolist()  # Extract metrics
        metrics_list[category] = metrics

    df = pd.read_excel(file_path, sheet_name="metrics")
    metrics_data = {}

    for index, row in df.iterrows():
        metric_name = row['Metric']
        description = row['Definition']
        criteria = []

        for col in df.columns[3:]:
            value = row[col]
            if pd.isna(value):
                break
            criteria.append(value)

        metrics_data[metric_name] = {
            'description': description,
            'criteria': criteria
        }

    return metrics_list, metrics_data


# Load the metrics and KPI data
file_path = resource_path('Master health check icsa Feb 24.xlsx')
metrics_list, metrics_data = load_metrics_from_excel(file_path)


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        selected_metrics = request.form.getlist('selected_metrics')
        if selected_metrics and isinstance(selected_metrics[0], str):
            try:
                selected_metrics = json.loads(selected_metrics[0])
            except json.JSONDecodeError:
                print("Error parsing the metrics list string.")

        return render_template('results.html', selected_metrics=selected_metrics, metrics_data=metrics_data)

    return render_template('index.html', metrics_list=metrics_list, metrics_data=metrics_data)


@app.route('/save_results', methods=['POST'])
def save_results():
    converted_input = {}
    user_input = request.form.to_dict()

    for key, value in user_input.items():
        try:
            if '.' in value:
                converted_input[key] = float(value)
            else:
                converted_input[key] = int(value)
        except (ValueError, TypeError):
            converted_input[key] = 'NA'

    calculate_outcome(converted_input)

    output_file = resource_path('Structured_Metrics_Output.xlsx')
    df = pd.read_excel(output_file)
    df = df.where(pd.notnull(df), '')

    results_data = df.to_dict(orient='records')
    return render_template('analysis.html', results_data=results_data)


@app.route('/close_excel', methods=['POST'])
def close_excel():
    try:
        force_close_excel()
        flash('Excel process has been successfully closed.')
    except Exception as e:
        flash(f"Error occurred while closing Excel: {e}")

    return redirect(url_for('index'))


if __name__ == '__main__':
    # Create a webview window and start it
    window = webview.create_window('DEMO APP', app, width=1500, height=800)
    webview.start()
