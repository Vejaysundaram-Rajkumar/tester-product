import openpyxl
import re
import xlwings as xw
import os

# Example user input with targets included
user_input = {
    'criteria_Average Days Delinquent_DSO': 1000,
    'criteria_Average Days Delinquent_BPDSO': 500,
    'criteria_Days of Inventory Outstanding_Average Inventory': 4000,
    'criteria_Days of Inventory Outstanding_Yearly Cost of Goods Solds (COGS)': 500,
    'criteria_Days Sales Outstanding_Average Account Receivables': 4000,
    'criteria_Days Sales Outstanding_Annual Sales': 200,
    'criteria_Cash Burn Rate_Cash Spent': 200,
    'criteria_Cash Burn Rate_Cash Received': 100,
    'criteria_Operating Cash Flow_Net Income': 1000,
    'criteria_Operating Cash Flow_Non Cash Expenses': 100,
    'criteria_Operating Cash Flow_Inc. in Working Capital': 20,
    'criteria_Days Payables Outstanding_Average Account Payables': 50,
    'criteria_Days Payables Outstanding_Yearly Cost Of Goods Solds (COGS)': 1,
    'criteria_Free Cash Flow_OCF': 100,
    'criteria_Free Cash Flow_Interest Payments': 50,
    'criteria_Free Cash Flow_Asset Purchase': 10,
    'criteria_Cash Conversion Cycle_DIO': 1000,
    'criteria_Cash Conversion Cycle_DSO': 500,
    'criteria_Cash Conversion Cycle_DPO': 100,
    'criteria_Overdues Ratio_Overdues': 500,
    'criteria_Overdues Ratio_Total Receivables': 50,
    'criteria_Cash Reserves in Days_Cash Reserves': 500,
    'criteria_Cash Reserves in Days_Average Daily Expenses': 5,
    'target_Cash Burn Rate': 100,
    'target_Operating Cash Flow': 1500,
    'target_Free Cash Flow': 200
}

# Step 1: Parse the user input to extract metrics, criteria, and targets
metrics_data = {}
targets_data = {}

for key, value in user_input.items():
    if key.startswith('criteria_'):
        # Extract the metric and criteria from the key
        pattern = r'criteria_(.*?)_(.*)'  # Regex to match the metric and criteria
        match = re.match(pattern, key)
        
        if match:
            metric_name = match.group(1)  # Extracted metric name
            criteria_name = match.group(2)  # Extracted criteria name

            # If the metric is not yet in the dictionary, initialize it with empty lists
            if metric_name not in metrics_data:
                metrics_data[metric_name] = [[], []]  # First list for criteria names, second for values

            # Add the criteria and its value to the lists
            metrics_data[metric_name][0].append(criteria_name)
            metrics_data[metric_name][1].append(value)
    
    elif key.startswith('target_'):
        # Extract the metric name for the target
        metric_name = key.replace('target_', '')
        targets_data[metric_name] = value

# Step 2: Load the Excel workbook and update the dataset with the extracted criteria values
file_path = 'Master health check icsa Feb 24.xlsx'
workbook = openpyxl.load_workbook(file_path, data_only=False)  # Load to keep formulas intact
sheet = workbook['metrics']  # Adjust to the correct sheet name if necessary

# Dictionary to store outcome by metric row for fetching later
metric_row_mapping = {}

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    metric_name = row[0].value  # Assuming metric name is in column A
    if not metric_name:
        continue  # Skip if there's no metric name

    # Check if the metric exists in the parsed user input
    if metric_name in metrics_data:
        criteria_list, value_list = metrics_data[metric_name]

        # Store the row number for outcome reference later
        metric_row_mapping[metric_name] = row[0].row  # Store the row number of this metric

        # Update the Excel sheet with user values corresponding to the criteria
        for idx, criteria in enumerate(criteria_list):
            for col_idx in range(3, 8):  # Columns D to H in the Excel sheet contain the criteria names
                if row[col_idx].value == criteria:
                    sheet.cell(row=row[0].row, column=col_idx + 1).value = value_list[idx]
                    break

# Save the updated workbook temporarily to recalculate outcomes
updated_file_path = 'Updated_Metrics_Dataset.xlsx'
workbook.save(updated_file_path)

# Step 3: Use xlwings to reopen the updated Excel file and fetch the computed outcomes
app = xw.App(visible=False)
workbook_xlwings = xw.Book(updated_file_path)
sheet_xlwings = workbook_xlwings.sheets['metrics']  # Adjust to the correct sheet name if necessary

# Step 4: Create a new workbook for the structured output
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.title = "Selected Metrics"

# Step 5: Add headers for the structured layout
headers = ["Metric", "Criteria", "Value", "Outcome", "Target"]
new_sheet.append(headers)

# Step 6: Loop over the metrics and write them into the new sheet in the desired structure
for metric_name, (criteria_list, value_list) in metrics_data.items():
    # Extract the outcome using xlwings
    metric_row_in_sheet = metric_row_mapping[metric_name]
    outcome_cell = sheet_xlwings.range(f'I{metric_row_in_sheet}').value  # Assuming outcome is in column I

    # Determine the target for this metric (if provided, else 'NA')
    target_value = targets_data.get(metric_name, 'NA')

    # Write the metric name, criteria, and values row by row
    for idx, criteria_name in enumerate(criteria_list):
        # If this is the first row for this metric, we write the outcome and target; otherwise, leave them blank
        if idx == 0:
            new_sheet.append([metric_name,  # Write the metric name once
                              criteria_name,  # Criteria name
                              value_list[idx],  # User input value
                              outcome_cell,  # Outcome
                              target_value])  # Target value
        else:
            new_sheet.append(["",  # Blank for metric name after first row
                              criteria_name,
                              value_list[idx],
                              "",  # Outcome blank for subsequent criteria rows
                              ""])  # Leave target blank for subsequent criteria rows

# Step 7: Close the xlwings workbook and save the new workbook with the structured output
workbook_xlwings.close()
app.quit()

# Step 8: Save the new workbook
new_file_path = 'Structured_Metrics_Output.xlsx'
new_workbook.save(new_file_path)

# Step 9: Delete the updated temporary Excel file
if os.path.exists(updated_file_path):
    os.remove(updated_file_path)

print(f"New structured Excel file saved as {new_file_path} and temporary file deleted.")
